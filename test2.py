import os
import csv
from openpyxl import load_workbook
import re

# Dossier d'entrée et de sortie
input_folder = 'input'
input_converted_folder = 'converted_input_files'
tecdoc_converted_folder = 'converted_tecdoc_file'
tecdoc_folder = 'ExempleTD'
tecdoc_file = 'TecDoc.xlsx'
mip_folder = 'MIP'
mip_converted_folder = 'MIP_converted'
mip_file = 'MIP.xlsx'
fusion_folder = 'fusion'
mip_done_folder = 'MIP_done'
mip_converted_file = 'mip.csv'

# Créer les dossiers de sortie s'ils n'existent pas
os.makedirs(input_converted_folder, exist_ok=True)
os.makedirs(tecdoc_converted_folder, exist_ok=True)
os.makedirs(mip_converted_folder, exist_ok=True)
os.makedirs(fusion_folder, exist_ok=True)
os.makedirs(mip_done_folder, exist_ok=True)

# Conversion des fichiers d'input en CSV
for input_file in os.listdir(input_folder):
    if input_file.endswith('.xlsx') and input_file != tecdoc_file:
        input_path = os.path.join(input_folder, input_file)
        converted_file = os.path.splitext(input_file)[0] + '.csv'
        converted_path = os.path.join(input_converted_folder, converted_file)

        # Charger le fichier xlsx
        wb = load_workbook(input_path)
        sheet = wb.active

        # Lire les données et les écrire dans le fichier CSV
        with open(converted_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Le fichier '{input_file}' a été converti en CSV : {converted_file}")

# Conversion du fichier TecDoc en CSV
tecdoc_path = os.path.join(tecdoc_folder, tecdoc_file)
converted_tecdoc_file = 'TecDoc.csv'
converted_tecdoc_path = os.path.join(tecdoc_converted_folder, converted_tecdoc_file)

# Charger le fichier xlsx TecDoc
tecdoc_wb = load_workbook(tecdoc_path)
tecdoc_sheet = tecdoc_wb.active

# Lire les en-têtes du fichier TecDoc
tecdoc_headers = [cell.value for cell in tecdoc_sheet[1]]

# Lire les données et les écrire dans le fichier CSV TecDoc
with open(converted_tecdoc_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(tecdoc_headers)
    for row in tecdoc_sheet.iter_rows(min_row=2, values_only=True):
        writer.writerow(row)

print(f"Le fichier '{tecdoc_file}' a été converti en CSV : {converted_tecdoc_file}")

# Conversion du fichier MIP en CSV
mip_path = os.path.join(mip_folder, mip_file)
mip_converted_file = 'mip.csv'
converted_mip_path = os.path.join(mip_converted_folder, mip_converted_file)

# Charger le fichier xlsx MIP
mip_wb = load_workbook(mip_path)
mip_sheet = mip_wb.active

# Lire les en-têtes du fichier MIP
mip_headers = [cell.value for cell in mip_sheet[1]]

# Lire les données et les écrire dans le fichier CSV MIP
with open(converted_mip_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(mip_headers)
    for row in mip_sheet.iter_rows(min_row=2, values_only=True):
        writer.writerow(row)

print(f"Le fichier '{mip_file}' a été converti en CSV : {mip_converted_file}")

# Dictionnaire des clés de jointure pour chaque fichier d'input
input_keys = {
    'valeo.csv': 'MPN (or OE)',
    'test.csv': 'code de fuuuuusion',
    # Ajoutez d'autres fichiers d'input avec leurs clés de jointure correspondantes ici
}

# Fusion des fichiers d'entrée avec TecDoc.csv
tecdoc_csv_path = os.path.join(tecdoc_converted_folder, converted_tecdoc_file)

for input_file, input_key in input_keys.items():
    input_path = os.path.join(input_converted_folder, input_file)
    fusion_file = f'fusion-{input_file}'

    with open(input_path, 'r', newline='', encoding='utf-8') as input_csv, \
            open(os.path.join(fusion_folder, fusion_file), 'w', newline='', encoding='utf-8') as fusion_csv:

        tecdoc_rows = []
        input_rows = []

        tecdoc_header = []
        input_header = []

        tecdoc_data = csv.reader(open(tecdoc_csv_path, 'r', newline='', encoding='utf-8'))
        input_data = csv.reader(input_csv)

        # Lire les en-têtes
        tecdoc_header = next(tecdoc_data)
        input_header = next(input_data)

        # Recherche de l'index de la clé de fusion dans les en-têtes
        tecdoc_key_index = tecdoc_header.index('MPN')
        input_key_index = input_header.index(input_key)

        # Lire les données dans des listes
        for tecdoc_row in tecdoc_data:
            tecdoc_rows.append(tecdoc_row)
        for input_row in input_data:
            input_rows.append(input_row)

        # Fusionner les données en utilisant la clé de fusion
        merged_rows = []
        for tecdoc_row in tecdoc_rows:
            for input_row in input_rows:
                if tecdoc_row[tecdoc_key_index] == input_row[input_key_index]:
                    merged_row = tecdoc_row + input_row
                    merged_rows.append(merged_row)

        # Écrire les données fusionnées dans le fichier de fusion
        fusion_writer = csv.writer(fusion_csv)
        fusion_writer.writerow(tecdoc_header + input_header)
        fusion_writer.writerows(merged_rows)

    print(f"La fusion des fichiers TecDoc.csv et {input_file} a été réalisée : {fusion_file}")

print("La fusion des fichiers a été réalisée avec succès.")

# écrire la data dans les fichiers MIP-{fusion file}.csv
mip_converted_path = os.path.join(mip_converted_folder, mip_converted_file)

def get_unmapped_fields(field_mapping, fusion_folder):
    mapped_fields = set()
    for mapping in field_mapping.values():
        mapped_fields.update(mapping.values())

    unmapped_fields = []

    for fusion_file in os.listdir(fusion_folder):
        with open(os.path.join(fusion_folder, fusion_file), 'r', encoding='utf-8-sig') as file:
            headers = file.readline().strip().split(',')
        unmapped_fields.extend(field for field in headers if field not in mapped_fields)

    return unmapped_fields

def write_mip_headers(field_mapping, num_attribute_fields, unmapped_fields):
    # Récupérer tous les headers existants dans MIP.csv
    with open(mip_converted_path, 'r', newline='', encoding='utf-8') as mip_csv:
        reader = csv.DictReader(mip_csv)
        mip_headers = reader.fieldnames

    # Récupérer les champs non mappés
    missing_fields = set(mip_headers) - set(field_mapping.keys())
    new_unmapped_fields = list(missing_fields.union(unmapped_fields))

    # Limiter le nombre de champs non mappés au maximum de 20
    max_unmapped_fields = 20 - num_attribute_fields
    if len(new_unmapped_fields) > max_unmapped_fields:
        new_unmapped_fields = new_unmapped_fields[:max_unmapped_fields]

    # Mettre à jour les headers de mip_headers avec les nouveaux champs créés et les champs non mappés
    mip_headers += list(missing_fields) + [f'Attribute Name {i}' for i in range(num_attribute_fields, num_attribute_fields + len(new_unmapped_fields))] + [f'Attribute Value {i}' for i in range(num_attribute_fields, num_attribute_fields + len(new_unmapped_fields))]

    new_field_mapping = {}
    for fusion_file in os.listdir(fusion_folder):
        if fusion_file.startswith('fusion-') and fusion_file.endswith('.csv'):
            mip_fusion_file = f'MIP-{fusion_file}'
            mip_fusion_path = os.path.join(mip_done_folder, mip_fusion_file)
            with open(mip_fusion_path, 'w', newline='', encoding='utf-8') as mip_fusion_csv:
                writer = csv.writer(mip_fusion_csv)
                fusion_fieldnames = list(mip_headers) + list(missing_fields)
                writer.writerow(fusion_fieldnames)

                # Mettre à jour le field_mapping avec les nouveaux champs créés
                new_field_mapping[fusion_file] = {}
                for field in missing_fields:
                    if field == 'Ktype':
                        new_field_mapping[fusion_file][field] = f'Compatible Product {num_compatible_product}'
                        num_compatible_product += 1
                    else:
                        new_field_mapping[fusion_file][field] = f'Attribute Name {num_attribute_fields}'
                        new_field_mapping[fusion_file][f'{field}_value'] = f'Attribute Value {num_attribute_fields}'
                        num_attribute_fields += 1

    field_mapping.update(new_field_mapping)

    print("Les headers des fichiers MIP ont été mis à jour avec succès.")
    return mip_headers, new_unmapped_fields

def process_data(fusion_folder, field_mapping, mip_headers):
    for fusion_file in os.listdir(fusion_folder):
        if fusion_file.startswith('fusion-') and fusion_file.endswith('.csv'):
            mip_fusion_file = f'MIP-{fusion_file}'
            mip_fusion_path = os.path.join(mip_done_folder, mip_fusion_file)
            with open(os.path.join(fusion_folder, fusion_file), 'r', newline='', encoding='utf-8') as fusion_csv, \
                    open(mip_fusion_path, 'w', newline='', encoding='utf-8') as mip_fusion_csv:
                fusion_reader = csv.DictReader(fusion_csv)
                mip_fusion_writer = csv.DictWriter(mip_fusion_csv, fieldnames=mip_headers)
                mip_fusion_writer.writeheader()

                mip_fusion_writer = csv.DictWriter(mip_fusion_csv, fieldnames=mip_headers)
                
                for row in fusion_reader:
                    mip_row = {}

                    # Écrire les champs fixes
                    mip_row['Channel ID'] = 'EBAY_FR'
                    mip_row['Condition'] = 'NEW'
                    mip_row['VAT Percent'] = 20
                    mip_row['Include eBay Product Details'] = False

                    # Écrire les champs mappés avec field_mapping
                    for fusion_field, mip_field in field_mapping[fusion_file].items():
                        if fusion_field in row:
                            mip_row[mip_field] = row[fusion_field]

                    # Écrire le champ 'Attribute Name 1' et le champ 'Attribute Value 1' de MIP-fusion_file avec le champ 'OE' de fusion_file
                    if 'OE' in field_mapping[fusion_file] and 'OE' in row:
                        oe_value = row['OE']
                        mip_row['Attribute Value 1'] = ' | '.join(oe_value.split(','))
                        mip_row['Attribute Name 1'] = 'OE'

                    # Écrire le champ "Compatible Product 1" de MIP-fusion_file avec le premier numéro récupéré du champ "Ktype" de fusion_file
                    ktype_field = field_mapping[fusion_file].get('Ktype')
                    if ktype_field and ktype_field in row:
                        ktype_value = row[ktype_field]
                        ktype_numbers = ktype_value.split(',')  # Remplacez ',' par le séparateur réel dans fusion_file
                        mip_row[ktype_field] = f"Ktype={ktype_numbers[0].strip()}" if ktype_numbers else ''

                    # Ajouter les champs non mappés de fusion_file à mip_row
                    for field in mip_headers:
                        if field.startswith('Attribute Name ') and field not in mip_row:
                            attribute_num = field.split(' ')[-1]
                            mip_row[field] = ''
                            if f'Attribute Value {attribute_num}' in mip_headers:
                                mip_row[f'Attribute Value {attribute_num}'] = ''

                    for field in row:
                        if field not in mip_row:
                            attribute_num = len(mip_headers) - 1
                            mip_row[f'Attribute Name {attribute_num}'] = field
                            mip_row[f'Attribute Value {attribute_num}'] = row[field]

                    mip_fusion_writer.writerow(mip_row)

    print("Les fichiers de fusion ont été traités avec succès.")


# Définir le field_mapping
field_mapping = {
    'fusion-valeo.csv': {
        "*ShippingProfileName": "Shipping policy",
        "*PaymentProfileName": "Payment Policy",
        "*ReturnProfileName": "Return Policy",
        "*StartPrice": "List Price",
        "StoreCategory": "Store Category Name 1"
        # Ajoutez d'autres mappages de champs ici
    },
    'fusion-test.csv': {
        "*ShippingProfileName": "Shipping policy",
        "*PaymentProfileName": "Payment Policy",
        "*ReturnProfileName": "Return Policy",
        "*StartPrice": "List Price",
        "StoreCategory": "Store Category Name 1"
        # Ajoutez d'autres mappages de champs ici
    },
    # Ajoutez d'autres mappages de champs pour les autres fichiers de fusion
}

num_compatible_product = 1
num_attribute_fields = 2

# Obtenir les champs non mappés
unmapped_fields = get_unmapped_fields(field_mapping, fusion_folder)

# Récupérer les headers MIP et mettre à jour les champs non mappés
mip_headers, unmapped_fields = write_mip_headers(field_mapping, num_attribute_fields, unmapped_fields)

# Traiter les fichiers de fusion
process_data(fusion_folder, field_mapping, mip_headers)

print("Le traitement des fichiers MIP a été réalisé avec succès.")