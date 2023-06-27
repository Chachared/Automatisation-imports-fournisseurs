import os
import csv
from openpyxl import load_workbook
import re

# Dossier d'entrée et de sortie
input_folder = 'input'
input_converted_folder = 'converted_input_files'
tecdoc_converted_folder = 'converted_tecdoc_file'
tecdoc_folder = 'ExempleTD'
tecdoc_file = 'TecDoc.xlsx'
mip_folder = 'MIP'
mip_converted_folder = 'MIP_converted'
mip_file = 'MIP.xlsx'
fusion_folder = 'fusion'
mip_done_folder = 'MIP_done'
mip_converted_file = 'mip.csv'

# Créer les dossiers de sortie s'ils n'existent pas
os.makedirs(input_converted_folder, exist_ok=True)
os.makedirs(tecdoc_converted_folder, exist_ok=True)
os.makedirs(mip_converted_folder, exist_ok=True)
os.makedirs(fusion_folder, exist_ok=True)
os.makedirs(mip_done_folder, exist_ok=True)

# Conversion des fichiers d'input en CSV
for input_file in os.listdir(input_folder):
    if input_file.endswith('.xlsx') and input_file != tecdoc_file:
        input_path = os.path.join(input_folder, input_file)
        converted_file = os.path.splitext(input_file)[0] + '.csv'
        converted_path = os.path.join(input_converted_folder, converted_file)

        # Charger le fichier xlsx
        wb = load_workbook(input_path)
        sheet = wb.active

        # Lire les données et les écrire dans le fichier CSV
        with open(converted_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Le fichier '{input_file}' a été converti en CSV : {converted_file}")

# Conversion du fichier TecDoc en CSV
tecdoc_path = os.path.join(tecdoc_folder, tecdoc_file)
converted_tecdoc_file = 'TecDoc.csv'
converted_tecdoc_path = os.path.join(tecdoc_converted_folder, converted_tecdoc_file)

# Charger le fichier xlsx TecDoc
tecdoc_wb = load_workbook(tecdoc_path)
tecdoc_sheet = tecdoc_wb.active

# Lire les en-têtes du fichier TecDoc
tecdoc_headers = [cell.value for cell in tecdoc_sheet[1]]

# Lire les données et les écrire dans le fichier CSV TecDoc
with open(converted_tecdoc_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(tecdoc_headers)
    for row in tecdoc_sheet.iter_rows(min_row=2, values_only=True):
        writer.writerow(row)

print(f"Le fichier '{tecdoc_file}' a été converti en CSV : {converted_tecdoc_file}")

# Conversion du fichier MIP en CSV
mip_path = os.path.join(mip_folder, mip_file)
mip_converted_file = 'mip.csv'
converted_mip_path = os.path.join(mip_converted_folder, mip_converted_file)

# Charger le fichier xlsx MIP
mip_wb = load_workbook(mip_path)
mip_sheet = mip_wb.active

# Lire les en-têtes du fichier MIP
mip_headers = [cell.value for cell in mip_sheet[1]]

# Lire les données et les écrire dans le fichier CSV MIP
with open(converted_mip_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(mip_headers)
    for row in mip_sheet.iter_rows(min_row=2, values_only=True):
        writer.writerow(row)

print(f"Le fichier '{mip_file}' a été converti en CSV : {mip_converted_file}")

# Dictionnaire des clés de jointure pour chaque fichier d'input
input_keys = {
    'valeo.csv': 'MPN (or OE)',
    'test.csv': 'code de fuuuuusion',
    # Ajoutez d'autres fichiers d'input avec leurs clés de jointure correspondantes ici
}

# Fusion des fichiers d'entrée avec TecDoc.csv
tecdoc_csv_path = os.path.join(tecdoc_converted_folder, converted_tecdoc_file)

for input_file, input_key in input_keys.items():
    input_path = os.path.join(input_converted_folder, input_file)
    fusion_file = f'fusion-{input_file}'

    with open(input_path, 'r', newline='', encoding='utf-8') as input_csv, \
            open(os.path.join(fusion_folder, fusion_file), 'w', newline='', encoding='utf-8') as fusion_csv:

        tecdoc_rows = []
        input_rows = []

        tecdoc_header = []
        input_header = []

        tecdoc_data = csv.reader(open(tecdoc_csv_path, 'r', newline='', encoding='utf-8'))
        input_data = csv.reader(input_csv)

        # Lire les en-têtes
        tecdoc_header = next(tecdoc_data)
        input_header = next(input_data)

        # Recherche de l'index de la clé de fusion dans les en-têtes
        tecdoc_key_index = tecdoc_header.index('MPN')
        input_key_index = input_header.index(input_key)

        # Lire les données dans des listes
        for tecdoc_row in tecdoc_data:
            tecdoc_rows.append(tecdoc_row)
        for input_row in input_data:
            input_rows.append(input_row)

        # Fusionner les données en utilisant la clé de fusion
        merged_rows = []
        for tecdoc_row in tecdoc_rows:
            for input_row in input_rows:
                if tecdoc_row[tecdoc_key_index] == input_row[input_key_index]:
                    merged_row = tecdoc_row + input_row
                    merged_rows.append(merged_row)

        # Écrire les données fusionnées dans le fichier de fusion
        fusion_writer = csv.writer(fusion_csv)
        fusion_writer.writerow(tecdoc_header + input_header)
        fusion_writer.writerows(merged_rows)

    print(f"La fusion des fichiers TecDoc.csv et {input_file} a été réalisée : {fusion_file}")

print("La fusion des fichiers a été réalisée avec succès.")


################## Ecriture des fichiers OUTPUT grace aux données des INPUT ###################################################

# Définir le field_mapping
field_mapping = {
    'fusion-valeo.csv': {
        "*ShippingProfileName": "Shipping policy",
        "*PaymentProfileName": "Payment Policy",
        "*ReturnProfileName": "Return Policy",
        "*StartPrice": "List Price",
        "StoreCategory": "Store Category Name 1",
        "Marque": "Attribute Value 1",
        "Type": "Attribute Value 2",
        "OE": "Attribute Value 3",
        "MPN": "Attribute Value 4",
        "Nombre de dents": "Attribute Value 5",
        "Force d'éjection (N)": "Attribute Value 6",
        "Poids": "Attribute Value 7",
        "Fixation de colonne de direction": "Attribute Value 8",
        "Diamètre du disque": "Attribute Value 9",
        "Info complementaire 1": "Additional Info",
        "SKU": "MPN_Brand",
        "Title" : "MPN_Brand_Type",
        # Ajoutez d'autres mappages de champs ici
    },
    'fusion-test.csv': {
        "*ShippingProfileName": "Shipping policy",
        "*PaymentProfileName": "Payment Policy",
        "*ReturnProfileName": "Return Policy",
        "*StartPrice": "List Price",
        "StoreCategory": "Store Category Name 1",
        "Marque": "Attribute Value 1",
        "Type": "Attribute Value 2",
        "OE": "Attribute Value 3",
        "MPN": "Attribute Value 4",
        "Nombre de dents": "Attribute Value 5",
        "Force d'éjection (N)": "Attribute Value 6",
        "Poids": "Attribute Value 7",
        "Fixation de colonne de direction": "Attribute Value 8",
        "Diamètre du disque": "Attribute Value 9",
        "Info complementaire 1": "Additional Info",
        "SKU": "MPN_Brand",
        "Title" : "MPN_Brand_Type",

        # Ajoutez d'autres mappages de champs ici
    },
    # Ajoutez d'autres mappages de champs pour les autres fichiers de fusion
}

# récupérer le template "MIP.csv"
mip_converted_path = os.path.join(mip_converted_folder, mip_converted_file)

separator_regex = re.compile(r'[,\\s;|]')

# Ecrire les fichiers MIP-fusion_file grâce aux fusion_files et au template MIP.csv
for fusion_file in os.listdir(fusion_folder):
    if fusion_file.startswith('fusion-') and fusion_file.endswith('.csv'):
        mip_fusion_file = f"MIP-{fusion_file}"
        mip_fusion_path = os.path.join(mip_done_folder, mip_fusion_file)

        with open(mip_converted_path, 'r', newline='', encoding='utf-8') as mip_converted_csv, \
                open(os.path.join(fusion_folder, fusion_file), 'r', newline='', encoding='utf-8') as fusion_csv, \
                open(mip_fusion_path, 'w', newline='', encoding='utf-8') as mip_fusion_csv:

            mip_converted_data = csv.reader(mip_converted_csv)
            fusion_data = csv.reader(fusion_csv)
            mip_fusion_writer = csv.writer(mip_fusion_csv)

            # Lire les en-têtes du fichier MIP converti
            mip_converted_headers = next(mip_converted_data)

            # Lire les en-têtes du fichier de fusion
            fusion_headers = next(fusion_data)

            # Créer un dictionnaire de mapping inversé pour les champs du fichier de fusion
            reverse_field_mapping = {v: k for k, v in field_mapping[fusion_file].items()}

            # Trouver les index des champs "Attribute Name" et "Attribute Value" existants dans le fichier MIP
            attribute_name_indices = [i for i, header in enumerate(mip_converted_headers) if header.startswith("Attribute Name")]
            attribute_value_indices = [i for i, header in enumerate(mip_converted_headers) if header.startswith("Attribute Value")]

            # Trouver le prochain index disponible pour les champs "Attribute Name" et "Attribute Value"
            next_attribute_name_index = len(attribute_name_indices) + 1
            next_attribute_value_index = len(attribute_value_indices) + 1

            mip_fusion_headers = []

            # Parcourir les en-têtes du fichier MIP converti
            for header in mip_converted_headers:
                cleaned_header = header.strip()  # Supprimer les espaces blancs au début et à la fin de l'en-tête
                if cleaned_header and cleaned_header != ",":
                    mip_fusion_headers.append(cleaned_header)

            # Parcourir les colonnes du fichier MIP converti
            for i in range(1, len(mip_converted_headers) + 1):
                attribute_name_field = f"Attribute Name {i}"
                attribute_value_field = f"Attribute Value {i}"
                if attribute_value_field in reverse_field_mapping:
                    attribute_name = reverse_field_mapping[attribute_value_field]  # Récupérer le nom du champ mappé
                    if attribute_name_field not in mip_fusion_headers:
                        mip_fusion_headers.append(attribute_name_field)  # Ajouter le champ "Attribute Name" dans les en-têtes
                        mip_fusion_headers.append(attribute_value_field)  # Ajouter le champ "Attribute Value" dans les en-têtes

            # Ajouter les en-têtes restants après le dernier champ "Attribute Value"
            mip_fusion_headers.extend(header for header in mip_converted_headers if header not in mip_fusion_headers)

            mip_fusion_writer.writerow(list(mip_fusion_headers))

            # Créer un dictionnaire de mappage inversé pour les champs du fichier de fusion
            reverse_mapping = {v: k for k, v in field_mapping[fusion_file].items()}

            # Parcourir les lignes du fichier de fusion
            for fusion_row in fusion_data:
                # Créer une liste pour chaque ligne du fichier de fusion MIP
                mip_fusion_row = []

                # Parcourir les colonnes du fichier MIP converti
                for i in range(len(mip_fusion_headers)):
                    header = mip_fusion_headers[i]
                    # Ecrire les champs Attribute Name
                    if header.startswith("Attribute Name"):
                        attribute_name_index = int(header.split(" ")[-1]) - 1
                        if attribute_name_index < len(reverse_mapping):
                            attribute_value_field = f"Attribute Value {attribute_name_index + 1}"
                            if attribute_value_field in reverse_mapping:
                                attribute_name = reverse_mapping[attribute_value_field]
                                mip_fusion_row.append(attribute_name)
                            else:
                                mip_fusion_row.append('')
                        else:
                            mip_fusion_row.append('')
                    # Ecrire les champs Attribute Value
                    elif header.startswith("Attribute Value"):
                        attribute_value_index = int(header.split(" ")[-1]) - 1
                        if attribute_value_index < len(reverse_mapping):
                            attribute_value_field = f"Attribute Value {attribute_value_index + 1}"
                            if attribute_value_field in reverse_mapping:
                                attribute_value = fusion_row[fusion_headers.index(reverse_mapping[attribute_value_field])]
                                mip_fusion_row.append(attribute_value)
                            else:
                                mip_fusion_row.append('')
                        else:
                            mip_fusion_row.append('')
                    # Récupérer la valeur du champ SKU à partir du mapping
                    elif header.startswith('SKU'):
                        sku_mapping = field_mapping[fusion_file].get("SKU")
                        if sku_mapping is not None:
                            sku_parts = sku_mapping.split("_")
                            sku_values = []
                            for part in sku_parts:
                                if part in fusion_headers:
                                    part_index = fusion_headers.index(part)
                                    part_value = fusion_row[part_index]
                                    sku_values.append(part_value)
                            sku_value = "_".join(sku_values)
                            mip_fusion_row.append(sku_value)
                    # Récupérer la valeur du champ Title à partir du mapping
                    elif header == "Title":
                        title_mapping = field_mapping[fusion_file].get("Title")
                        if title_mapping is not None:
                            title_parts = title_mapping.split("_")
                            title_values = []
                            for part in title_parts:
                                if part in fusion_headers:
                                    part_index = fusion_headers.index(part)
                                    part_value = fusion_row[part_index]
                                    title_values.append(part_value)
                            title_value = " ".join(title_values)
                            mip_fusion_row.append(title_value)
                    # Ecrire les champs à valeurs fixes
                    elif header == "Localized For":
                        mip_fusion_row.append('fr_FR')
                    elif header == "Subtitle":
                        mip_fusion_row.append('')  # Attribuer une chaîne vide au champ "Subtitle" (option payante)
                    elif header == "Condition":
                        mip_fusion_row.append('NEW')
                    elif header == "Channel ID":
                        mip_fusion_row.append('EBAY_FR')
                    elif header == "VAT Percent":
                        mip_fusion_row.append(20)
                    elif header  == "Include eBay Product Details":
                        mip_fusion_row.append("false")
                    # Ecrire tous les autres champs mappés
                    elif not header.startswith("Attribute Value") and not header.startswith("Attribute Name"):
                        if header in reverse_mapping:
                            value_field = reverse_mapping[header]
                            if value_field in fusion_headers:
                                value_index = fusion_headers.index(value_field)
                                value = fusion_row[value_index]
                                mip_fusion_row.append(value)
                            else:
                                mip_fusion_row.append('')
                        else:
                            mip_fusion_row.append('')

                mip_fusion_writer.writerow(mip_fusion_row)

        print(f"Le fichier de fusion {fusion_file} a été traité avec succès : {mip_fusion_file}")

print("Les fichiers de sortie ont été écrits avec succès.")

## todo 2 : séparer les numéros de Ktype dans le champ compatible product 1. Créer un nouveau Compatible Product {i} pour chaque numéro de Ktype trouvé. L'écrire "Ktype={Ktype}"
## todo 3 : séparer les numéros du champ Attribute value qui correspond à "OE" par des '|', au lieu des séparateurs actuels. Puis limiter le nombre de numéros à 65, et créer un nouveau champ Attribute Name/Value pour les numéros au dela de 65, et ainsi de suite
## todo 4 : ajouter le reste du mapping
## todo 5 : mapper les vrais fichiers (y compris pour fusion tecdoc/input)