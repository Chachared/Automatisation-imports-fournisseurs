import csv
from openpyxl import load_workbook

#########################convertir et fusionner les fichiers d'entrée (input, complément d'infos de tecdoc...)########################################
# convertir le fichier input depuis excel vers csv
valeo_file = load_workbook(filename = 'Input/valeo.xlsx')
sheet_valeo = valeo_file.active
csv_data_valeo = []
for value in sheet_valeo.iter_rows(values_only=True):
    csv_data_valeo.append(list(value))
with open('Valeo.csv', 'w') as valeo_csv:
    writer = csv.writer(valeo_csv, delimiter=',')
    for line in csv_data_valeo:
        writer.writerow(line)

# convertir le fichier tecdoc depuis excel vers csv
tecdoc_file = load_workbook(filename = 'ExempleTD/TecDoc.xlsx')
sheet_tecdoc = tecdoc_file.active
csv_data_tecdoc = []
for value in sheet_tecdoc.iter_rows(values_only=True):
    csv_data_tecdoc.append(list(value))
with open('Tecdoc.csv', 'w') as tecdoc_csv:
    writer = csv.writer(tecdoc_csv, delimiter=',')
    for line in csv_data_tecdoc:
        writer.writerow(line)

# lire les données dans les csv
valeo_data = []
with open('Valeo.csv', 'r') as valeo_csv:
    reader = csv.DictReader(valeo_csv)
    for row in reader:
        valeo_data.append(row)

tecdoc_data = []
with open('Tecdoc.csv', 'r') as tecdoc_csv:
    reader = csv.DictReader(tecdoc_csv)
    for row in reader:
        tecdoc_data.append(row)

# écrire les données du csv tecdoc dans un dictionnaire
tecdoc_dict = {}
for row in tecdoc_data:
    mpn = row['MPN']
    tecdoc_dict[mpn] = row

# parcourir les données dans le csv valeo pour trouver le champ qui sert de clé de jointure
merged_data = []
for row in valeo_data:
    mpn = row['MPN (or OE)']
    if mpn in tecdoc_dict:
        merged_row = {**row, **tecdoc_dict[mpn]}
        merged_data.append(merged_row)

# fusionner les 2 csv sur le champ "MPN" / "MPN (or OE)"
fieldnames = list(merged_data[0].keys())

with open('fusion-input-tecdoc.csv', 'w', newline='') as fusion_csv:
    writer = csv.DictWriter(fusion_csv, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(merged_data)

#####################écrire les datas collectées en entrée vers la sortie (fichier output MIP ebay)############################################

# convertir le fichier MIP en csv
MIP_file = load_workbook(filename = 'MIP.xlsx')
sheet_MIP = MIP_file.active
csv_data_MIP = []
for value in sheet_MIP.iter_rows(values_only=True):
    csv_data_MIP.append(list(value))
with open('MIP.csv', 'w') as MIP_csv:
    writer = csv.writer(MIP_csv, delimiter=',')
    for line in csv_data_MIP:
        writer.writerow(line)

# lire les données dans le csv fusion
fusion_data = []
with open('fusion-input-tecdoc.csv', 'r') as fusion_csv:
    reader = csv.DictReader(fusion_csv)
    for row in reader:
        fusion_data.append(row)

# Lire les noms de champs du fichier MIP
with open('MIP.csv', 'r') as MIP_csv:
    reader = csv.DictReader(MIP_csv)
    mip_fieldnames = reader.fieldnames

# Mapper les champs correspondants entre fusion et MIP
mapping = {
    'MPN': 'MPN',
    'List Price': '*StartPrice',
    'Total Ship To Home Quantity': '*Quantity',
    'Shipping policy': '*ShippingProfileName',
    'Return Policy': '*ReturnProfileName',
    'Payment Policy': '*PaymentProfileName'
}

# Écrire les datas mappées dans une nouvelle liste
mapped_data = []
for row in fusion_data:
    mapped_row = {}
    for mip_field in mip_fieldnames:
        if mip_field in mapping:
            mapped_row[mip_field] = row[mapping[mip_field]]
        else:
            mapped_row[mip_field] = ''  # Champ vide si aucune donnée correspondante
    mapped_data.append(mapped_row)

# Écrire les données mappées dans le fichier MIP
with open('MIP.csv', 'w', newline='') as mip_csv:
    writer = csv.DictWriter(mip_csv, fieldnames=mip_fieldnames)
    writer.writeheader()
    writer.writerows(mapped_data)