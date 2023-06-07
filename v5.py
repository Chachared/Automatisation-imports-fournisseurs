import os
import csv
from openpyxl import load_workbook

# Dossier d'entrée et de sortie
input_folder = 'input'
input_converted_folder = 'converted_input_files'
tecdoc_converted_folder = 'converted_tecdoc_file'
tecdoc_folder = 'ExempleTD'
tecdoc_file = 'TecDoc.xlsx'
mip_folder = 'MIP'
mip_converted_folder = 'MIP_converted'
mip_file = 'MIP.xlsx'
fusion_folder = 'fusion'
mip_done_folder = 'MIP_done'
mip_converted_file = 'mip.csv'

# Créer les dossiers de sortie s'ils n'existent pas
os.makedirs(input_converted_folder, exist_ok=True)
os.makedirs(tecdoc_converted_folder, exist_ok=True)
os.makedirs(mip_converted_folder, exist_ok=True)
os.makedirs(fusion_folder, exist_ok=True)
os.makedirs(mip_done_folder, exist_ok=True)

# Conversion des fichiers d'input en CSV
for input_file in os.listdir(input_folder):
    if input_file.endswith('.xlsx') and input_file != tecdoc_file:
        input_path = os.path.join(input_folder, input_file)
        converted_file = os.path.splitext(input_file)[0] + '.csv'
        converted_path = os.path.join(input_converted_folder, converted_file)

        # Charger le fichier xlsx
        wb = load_workbook(input_path)
        sheet = wb.active

        # Lire les données et les écrire dans le fichier CSV
        with open(converted_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Le fichier '{input_file}' a été converti en CSV : {converted_file}")

# Conversion du fichier TecDoc en CSV
tecdoc_path = os.path.join(tecdoc_folder, tecdoc_file)
converted_tecdoc_file = 'TecDoc.csv'
converted_tecdoc_path = os.path.join(tecdoc_converted_folder, converted_tecdoc_file)

# Charger le fichier xlsx TecDoc
tecdoc_wb = load_workbook(tecdoc_path)
tecdoc_sheet = tecdoc_wb.active

# Lire les en-têtes du fichier TecDoc
tecdoc_headers = [cell.value for cell in tecdoc_sheet[1]]

# Lire les données et les écrire dans le fichier CSV TecDoc
with open(converted_tecdoc_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(tecdoc_headers)
    for row in tecdoc_sheet.iter_rows(min_row=2, values_only=True):
        writer.writerow(row)

print(f"Le fichier '{tecdoc_file}' a été converti en CSV : {converted_tecdoc_file}")

# Conversion du fichier MIP en CSV
mip_path = os.path.join(mip_folder, mip_file)
converted_mip_file = 'mip.csv'
converted_mip_path = os.path.join(mip_converted_folder, converted_mip_file)

# Charger le fichier xlsx MIP
mip_wb = load_workbook(mip_path)
mip_sheet = mip_wb.active

# Lire les en-têtes du fichier MIP
mip_headers = [cell.value for cell in mip_sheet[1]]

# Lire les données et les écrire dans le fichier CSV MIP
with open(converted_mip_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(mip_headers)
    for row in mip_sheet.iter_rows(min_row=2, values_only=True):
        writer.writerow(row)

print(f"Le fichier '{mip_file}' a été converti en CSV : {converted_mip_file}")

# Dictionnaire des clés de jointure pour chaque fichier d'input
input_keys = {
    'valeo.csv': 'MPN (or OE)',
    'test.csv': 'code de fuuuuusion',
    # Ajoutez d'autres fichiers d'input avec leurs clés de jointure correspondantes ici
}

# Fusion des fichiers d'entrée avec TecDoc.csv
tecdoc_csv_path = os.path.join(tecdoc_converted_folder, converted_tecdoc_file)

for input_file, input_key in input_keys.items():
    input_path = os.path.join(input_converted_folder, input_file)
    fusion_file = f'fusion-{input_file}'

    with open(input_path, 'r', newline='', encoding='utf-8') as input_csv, \
            open(os.path.join(fusion_folder, fusion_file), 'w', newline='', encoding='utf-8') as fusion_csv:

        tecdoc_rows = []
        input_rows = []

        tecdoc_header = []
        input_header = []

        tecdoc_data = csv.reader(open(tecdoc_csv_path, 'r', newline='', encoding='utf-8'))
        input_data = csv.reader(input_csv)

        # Lire les en-têtes
        tecdoc_header = next(tecdoc_data)
        input_header = next(input_data)

        # Recherche de l'index de la clé de fusion dans les en-têtes
        tecdoc_key_index = tecdoc_header.index('MPN')
        input_key_index = input_header.index(input_key)

        # Lire les données dans des listes
        for tecdoc_row in tecdoc_data:
            tecdoc_rows.append(tecdoc_row)
        for input_row in input_data:
            input_rows.append(input_row)

        # Fusionner les données en utilisant la clé de fusion
        merged_rows = []
        for tecdoc_row in tecdoc_rows:
            for input_row in input_rows:
                if tecdoc_row[tecdoc_key_index] == input_row[input_key_index]:
                    merged_row = tecdoc_row + input_row
                    merged_rows.append(merged_row)

        # Écrire les données fusionnées dans le fichier de fusion
        fusion_writer = csv.writer(fusion_csv)
        fusion_writer.writerow(tecdoc_header + input_header)
        fusion_writer.writerows(merged_rows)

    print(f"La fusion des fichiers TecDoc.csv et {input_file} a été réalisée : {fusion_file}")

print("La fusion des fichiers a été réalisée avec succès.")

# écrire la data dans les fichiers MIP-{fusion file}.csv
mip_converted_path = os.path.join(mip_converted_folder, mip_converted_file)

# Écrire les en-têtes de MIP.csv dans les fichiers MIP fusionnés
def write_mip_headers():
    mip_fieldnames = []
    with open(mip_converted_path, 'r') as mip_csv:
        mip_reader = csv.DictReader(mip_csv)
        mip_fieldnames = mip_reader.fieldnames

    mip_files = [f for f in os.listdir(mip_done_folder) if f.startswith("MIP-")]
    for mip_file in mip_files:
        mip_path = os.path.join(mip_done_folder, mip_file)
        with open(mip_path, 'w', newline='') as mip_csv:
            mip_writer = csv.DictWriter(mip_csv, fieldnames=mip_fieldnames)
            mip_writer.writeheader()

# Écrire la data dans les fichiers MIP fusionnés
def process_data(fusion_folder, field_mapping):
    fusion_files = os.listdir(fusion_folder)

    for fusion_file in fusion_files:
        if not fusion_file.endswith('.csv'):
            continue

        fusion_path = os.path.join(fusion_folder, fusion_file)
        mip_file = fusion_file.replace("fusion-", "MIP-fusion-")
        mip_path = os.path.join(mip_done_folder, mip_file)

        if not os.path.isfile(fusion_path):
            continue

        mapping = field_mapping.get(fusion_file, {})

        mip_fieldnames = []
        with open(mip_converted_path, 'r') as mip_csv:
            mip_reader = csv.DictReader(mip_csv)
            mip_fieldnames = mip_reader.fieldnames

        with open(fusion_path, 'r') as fusion_csv, open(mip_path, 'w', newline='') as mip_csv:
            fusion_reader = csv.DictReader(fusion_csv)

            mip_writer = csv.DictWriter(mip_csv, fieldnames=mip_fieldnames)
            mip_writer.writeheader()

            for row in fusion_reader:
                mip_row = {}

                for fusion_field, mip_field in mapping.items():
                    if fusion_field in row:
                        mip_row[mip_field] = row[fusion_field]
                    else:
                        mip_row[mip_field] = ''

                mip_writer.writerow(mip_row)



# Dictionnaire de mapping des champs entre les fichiers de fusion et MIP
field_mapping = {
    'fusion-valeo.csv': {
        'Marque': 'Attribute Value 1',
        'Type': 'Attribute Value 2',
        'Nombre de dents': 'Attribute Value 3',
        'Force d\'éjection (N)': 'Attribute Value 4',
        'Poids': 'Attribute Value 5',
        'Fixation de colonne de direction': 'Attribute Value 6',
        'Diamètre du disque': 'Attribute Value 7',
        'Info complementaire 1': 'Attribute Value 8',
        'OE': 'Attribute Value 9',
    },
    'fusion-test.csv': {
        'Marque': 'Attribute Value 1',
        'Type': 'Attribute Value 2',
        'Nombre de dents': 'Attribute Value 3',
        'Force d\'éjection (N)': 'Attribute Value 4',
        'Poids': 'Attribute Value 5',
        'Fixation de colonne de direction': 'Attribute Value 6',
        'Diamètre du disque': 'Attribute Value 7',
        'Info complementaire 1': 'Attribute Value 8',
        'OE': 'Attribute Value 9',
        # Ajoutez d'autres mappages de champs ici
    },
    # Ajoutez d'autres mappages de champs pour les autres fichiers de fusion
}

# Parcourir les fichiers de fusion et les traiter en masse
process_data(fusion_folder, field_mapping)

print("L'écriture des fichiers MIP par fournisseur est terminée")


##n'écrit pas tous les champs MIP dans les fichiers MIP-fusion => à résoudre