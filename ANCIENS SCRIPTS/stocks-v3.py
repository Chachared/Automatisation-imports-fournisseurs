import csv
from openpyxl import load_workbook

#########################convertir et fusionner les fichiers d'entrée (input, complément d'infos de tecdoc...)########################################
# convertir le fichier input depuis excel vers csv
valeo_file = load_workbook(filename = 'Input/valeo.xlsx')
sheet_valeo = valeo_file.active
csv_data_valeo = []
for value in sheet_valeo.iter_rows(values_only=True):
    csv_data_valeo.append(list(value))
with open('Valeo.csv', 'w') as valeo_csv:
    writer = csv.writer(valeo_csv, delimiter=',')
    for line in csv_data_valeo:
        writer.writerow(line)

# convertir le fichier tecdoc depuis excel vers csv
tecdoc_file = load_workbook(filename = 'ExempleTD/TecDoc.xlsx')
sheet_tecdoc = tecdoc_file.active
csv_data_tecdoc = []
for value in sheet_tecdoc.iter_rows(values_only=True):
    csv_data_tecdoc.append(list(value))
with open('Tecdoc.csv', 'w') as tecdoc_csv:
    writer = csv.writer(tecdoc_csv, delimiter=',')
    for line in csv_data_tecdoc:
        writer.writerow(line)

# lire les données dans les csv
valeo_data = []
with open('Valeo.csv', 'r') as valeo_csv:
    reader = csv.DictReader(valeo_csv)
    for row in reader:
        valeo_data.append(row)

tecdoc_data = []
with open('Tecdoc.csv', 'r') as tecdoc_csv:
    reader = csv.DictReader(tecdoc_csv)
    for row in reader:
        tecdoc_data.append(row)

# écrire les données du csv tecdoc dans un dictionnaire
tecdoc_dict = {}
for row in tecdoc_data:
    mpn = row['MPN']
    tecdoc_dict[mpn] = row

# parcourir les données dans le csv valeo pour trouver le champ qui sert de clé de jointure
merged_data = []
for row in valeo_data:
    mpn = row['MPN (or OE)']
    if mpn in tecdoc_dict:
        merged_row = {**row, **tecdoc_dict[mpn]}
        merged_data.append(merged_row)

# fusionner les 2 csv sur le champ "MPN" / "MPN (or OE)"
fieldnames = list(merged_data[0].keys())

with open('fusion-input-tecdoc.csv', 'w', newline='') as fusion_csv:
    writer = csv.DictWriter(fusion_csv, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(merged_data)

#####################écrire les datas collectées en entrée vers la sortie (fichier output MIP ebay)############################################

# convertir le fichier MIP en csv
MIP_file = load_workbook(filename = 'MIP.xlsx')
sheet_MIP = MIP_file.active
csv_data_MIP = []
for value in sheet_MIP.iter_rows(values_only=True):
    csv_data_MIP.append(list(value))
with open('MIP.csv', 'w') as MIP_csv:
    writer = csv.writer(MIP_csv, delimiter=',')
    for line in csv_data_MIP:
        writer.writerow(line)

# lire les données dans le csv fusion
fusion_data = []
with open('fusion-input-tecdoc.csv', 'r') as fusion_csv:
    reader = csv.DictReader(fusion_csv)
    for row in reader:
        fusion_data.append(row)

# Lire les noms de champs du fichier MIP
with open('MIP.csv', 'r') as MIP_csv:
    reader = csv.DictReader(MIP_csv)
    mip_fieldnames = reader.fieldnames

# Mapper les champs correspondants entre fusion et MIP
mapping = {
    'MPN': 'MPN',
    'List Price': '*StartPrice',
    'Total Ship To Home Quantity': '*Quantity',
    'Shipping policy': '*ShippingProfileName',
    'Return Policy': '*ReturnProfileName',
    'Payment Policy': '*PaymentProfileName'
}

# Écrire les datas mappées dans une nouvelle liste
mapped_data = []
for row in fusion_data:
    mapped_row = {}
    for mip_field in mip_fieldnames:
        if mip_field in mapping:
            mapped_row[mip_field] = row[mapping[mip_field]]
        # construire la valeur du champ SKU en concaténant les valeurs des champs MPN et Brand
        elif mip_field == 'SKU':
            mapped_row[mip_field] = row['MPN'] + '_' + row['Brand']
        # remplir les champs Ktype en séparant les Ktypes renseignés dans fusion
        elif mip_field.startswith('Compatible Product'):
            compatible_product_number = mip_field.split()[-1]  # Récupérer le numéro de "Compatible Product"
            ktype_values = row['Ktype'].split(',')  # Diviser les valeurs du champ "Ktype" par ";"
            if int(compatible_product_number) <= len(ktype_values):
                mapped_row[mip_field] = 'Ktype=' + ktype_values[int(compatible_product_number) - 1]
            else:
                mapped_row[mip_field] = ''  # Champ vide si l'index dépasse le nombre de valeurs disponibles
        # remplir les champs d'attributs avec le header du champ dans le champ Attribute Name, et la valeur du champ dans Attribute Value
        elif mip_field == 'Attribute Name 1':
            mapped_row[mip_field] = 'Brand'
        elif mip_field == 'Attribute Value 1':
            mapped_row[mip_field] = row['Brand']
        elif mip_field == 'Attribute Name 2':
            mapped_row[mip_field] = 'Type'
        elif mip_field == 'Attribute Value 2':
            mapped_row[mip_field] = row['Type']
        elif mip_field == 'Attribute Name 3':
            mapped_row[mip_field] = 'Nombre de dents'
        elif mip_field == 'Attribute Value 3':
            mapped_row[mip_field] = row['Nombre de dents']
        elif mip_field == 'Attribute Name 4':
            mapped_row[mip_field] = 'Force d\'éjection (N)'
        elif mip_field == 'Attribute Value 4':
            mapped_row[mip_field] = row['Force d\'éjection (N)']
        elif mip_field == 'Attribute Name 5':
            mapped_row[mip_field] = 'Poids'
        elif mip_field == 'Attribute Value 5':
            mapped_row[mip_field] = row['Poids']
        elif mip_field == 'Attribute Name 6':
            mapped_row[mip_field] = 'Fixation de colonne de direction'
        elif mip_field == 'Attribute Value 6':
            mapped_row[mip_field] = row['Fixation de colonne de direction']
        elif mip_field == 'Attribute Name 7':
            mapped_row[mip_field] = 'Diamètre du disque'
        elif mip_field == 'Attribute Value 7':
            mapped_row[mip_field] = row['Diamètre du disque']
        elif mip_field == 'Attribute Name 8':
            mapped_row[mip_field] = 'Info complementaire 1'
        elif mip_field == 'Attribute Value 8':
            mapped_row[mip_field] = row['Info complementaire 1']
        # écrire les OE en remplaçant les virgules par des |
        elif mip_field == 'Attribute Name 9':
            mapped_row[mip_field] = 'OE'
        elif mip_field == 'Attribute Value 9':
            mapped_row[mip_field] = row['OE'].replace(',', ' | ') if 'OE' in row else ''
        else:
            mapped_row[mip_field] = ''  # Champ vide si aucune donnée correspondante
    mapped_data.append(mapped_row)


# Écrire les données mappées dans le fichier MIP
with open('MIP.csv', 'w', newline='') as mip_csv:
    writer = csv.DictWriter(mip_csv, fieldnames=mip_fieldnames)
    writer.writeheader()
    writer.writerows(mapped_data)