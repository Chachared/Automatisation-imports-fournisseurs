import os
import csv
from openpyxl import load_workbook

# Définir les champs MIP
mip_fieldnames = [
    'SKU',
    'Localized For',
    'Title',
    'Subtitle',
    'Product Description',
    'Additional Info',
    'Group Picture URL',
    'UPC',
    'ISBN',
    'EAN',
    'Condition',
    'Condition Description',
    'Total Ship To Home Quantity',
    'Channel ID',
    'Category',
    'Shipping policy',
    'Payment Policy',
    'Return Policy',
    'List Price',
    'VAT Percent',
    'Max Quantity Per Buyer',
    'Strikethrough Price',
    'Store Category Name 1',
    'Store Category Name 2',
    'Include eBay Product Details',
    'Template Name'
]

# Ajouter les champs picture url dans mip_fieldnames
num_pictures = 12
for i in range(1, num_pictures + 1):
    picture_url = f'Picture URL {i}'
    mip_fieldnames.append(picture_url)

# Ajouter les champs attribute dans mip_fieldnames
num_attributes = 25
for i in range(1, num_attributes + 1):
    attribute_name = f'Attribute Name {i}'
    attribute_value = f'Attribute Value {i}'
    mip_fieldnames.append(attribute_name)
    mip_fieldnames.append(attribute_value)

# Ajouter les champs compatible products dans mip_fieldnames
num_compatible_products = 176
for i in range(1, num_compatible_products + 1):
    compatible_product = f'Compatible Product {i}'
    mip_fieldnames.append(compatible_product)

# Mapper les champs correspondants entre fusion-input-data et MIP
mapping = {
    'MPN': 'MPN',
    'List Price': '*StartPrice',
    'Total Ship To Home Quantity': '*Quantity',
    'Shipping policy': '*ShippingProfileName',
    'Return Policy': '*ReturnProfileName',
    'Payment Policy': '*PaymentProfileName'
}

# Ajouter les champs numérotés au mapping
for i in range(1, num_pictures + 1):
    picture_url = f'Picture URL {i}'
    mapping[picture_url] = picture_url

for i in range(1, num_attributes + 1):
    attribute_name = f'Attribute Name {i}'
    attribute_value = f'Attribute Value {i}'
    mapping[attribute_name] = attribute_name
    mapping[attribute_value] = attribute_value

for i in range(1, num_compatible_products + 1):
    compatible_product = f'Compatible Product {i}'
    mapping[compatible_product] = compatible_product

# Champs Attribute Name et Attribute Value spécifiques à chaque fichier d'input
input_attribute_mapping = {
    'valeo.csv': [
        {'Attribute Name 1': 'Brand', 'Attribute Value 1': 'Brand'},
        {'Attribute Name 2': 'Type', 'Attribute Value 2': 'Type'},
        {'Attribute Name 3': 'Nombre de dents', 'Attribute Value 3': 'Nombre de dents'},
        {'Attribute Name 4': "Force d'éjection (N)", 'Attribute Value 4': "Force d'éjection (N)"},
        {'Attribute Name 5': 'Poids', 'Attribute Value 5': 'Poids'}
    ]
}

# Chemin d'accès au fichier TecDoc.xlsx
tecdoc_excel_file = 'ExempleTD/TecDoc.xlsx'

# Charger le fichier TecDoc.xlsx et le convertir en CSV
tecdoc_csv_file = 'tecdoc.csv'
wb = load_workbook(filename=tecdoc_excel_file)
ws = wb.active
with open(tecdoc_csv_file, 'w', newline='', encoding='utf-8') as csv_file:
    writer = csv.writer(csv_file)
    writer.writerows(ws.values)

# Charger le fichier tecdoc.csv
tecdoc_data = []
with open(tecdoc_csv_file, 'r', encoding='utf-8') as tecdoc_csv:
    reader = csv.DictReader(tecdoc_csv)
    for row in reader:
        tecdoc_data.append(row)

# Conversion des fichiers XLSX en CSV
input_folder = 'input'
for input_file in os.listdir(input_folder):
    if input_file.endswith('.xlsx'):
        input_file_path = os.path.join(input_folder, input_file)
        output_file = os.path.splitext(input_file)[0] + '.csv'
        output_file_path = os.path.join(input_folder, output_file)

        # Charger le fichier input XLSX
        input_workbook = load_workbook(filename=input_file_path)
        input_sheet = input_workbook.active

        # Convertir le fichier input XLSX en CSV
        with open(output_file_path, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerows(input_sheet.values)

# Fusionner les tables input et tecdoc
merged_data = []
for input_file in os.listdir(input_folder):
    if input_file.endswith('.csv'):
        input_filename = os.path.join(input_folder, input_file)
        input_name, _ = os.path.splitext(input_file)
        input_attributes = input_attribute_mapping.get(input_file, [])
        input_key_mapping = {
            'valeo.csv': 'MPN (or OE)',
            # 'autre_fichier.csv': 'Autre champ de jointure',
            # 'encore_un_autre_fichier.csv': 'Champ de jointure différent'
        }
        input_key = input_key_mapping.get(input_file, 'OE')

        # Charger le fichier input
        with open(input_filename, 'r', encoding='utf-8') as input_file:
            reader = csv.DictReader(input_file)
            input_data = list(reader)

        # Fusionner les données avec tecdoc_data
        for input_row in input_data:
            merged_row = {}
            for tecdoc_row in tecdoc_data:
                if input_row[input_key] == tecdoc_row['MPN']:
                    merged_row = tecdoc_row.copy()
                    break
            merged_row.update({mapping[k]: input_row[v] for d in input_attributes for k, v in d.items()})

            # Effectuer le mapping pour les champs spécifiques à chaque fichier d'entrée
            for mip_field, input_field_mapping in mapping.items():
                attribute_name = input_field_mapping.get('Attribute Name', '')
                attribute_value = input_field_mapping.get('Attribute Value', '')
                merged_row[mip_field] = input_row.get(attribute_value, '')

            merged_data.append(merged_row)

# Effectuer le mapping global pour les champs du fichier MIP
mapped_data = []
for row in merged_data:
    mapped_row = {}
    for mip_field in mip_fieldnames:
        if mip_field in mapping:
            mapped_row[mip_field] = row[mapping[mip_field]]
        elif mip_field == 'SKU':
            mapped_row[mip_field] = row['MPN'] + '_' + row['Brand']
        else:
            mapped_row[mip_field] = ''
    mapped_data.append(mapped_row)

# Écrire les données fusionnées dans le fichier fusion-input-data.csv
with open('fusion-input-data.csv', 'w', newline='', encoding='utf-8') as data_file:
    writer = csv.DictWriter(data_file, fieldnames=mip_fieldnames, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(mapped_data)

# Charger le fichier fusion-input-data.csv
mip_data = []
with open('fusion-input-data.csv', 'r', encoding='utf-8') as mip_csv:
    reader = csv.DictReader(mip_csv)
    for row in reader:
        mip_data.append(row)

# Écrire les données mappées dans le fichier MIP.csv
with open('MIP.csv', 'w', newline='', encoding='utf-8') as mip_file:
    writer = csv.DictWriter(mip_file, fieldnames=mip_fieldnames, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(mip_data)
