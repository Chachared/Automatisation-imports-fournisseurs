import os
import csv
from openpyxl import load_workbook

# Dossier d'entrée et de sortie
input_folder = 'Input'
converted_folder = 'converted_input_files'
fusion_folder = 'fusion_input'
mip_folder = 'MIP_done'
converted_tecdoc_folder = 'converted_tecdoc_files'
converted_mip_folder = 'MIP_converted'

# Créer les dossiers de sortie si nécessaire
os.makedirs(converted_folder, exist_ok=True)
os.makedirs(fusion_folder, exist_ok=True)
os.makedirs(mip_folder, exist_ok=True)
os.makedirs(converted_tecdoc_folder, exist_ok=True)

# Mapping générique pour les fichiers MIP
mip_mapping = {
    'MPN': 'MPN',
    'Attribute Value 1': 'Marque',
    'Attribute Value 2': 'Type',
    'Attribute Value 3': 'Nombre de dents',
    'Attribute Value 4': 'Force d\'éjection (N)',
    'Attribute Value 5': 'Poids',
    'Attribute Value 6': 'Fixation de colonne de direction',
    'Attribute Value 7': 'Diamètre du disque',
    'Attribute Value 8': 'Info complementaire 1',
    'Attribute Value 9': 'OE',
}

# Mapping générique pour les champs de fusion
header_mapping = {
    'Attribute Name 1': 'Marque',
    'Attribute Name 2': 'Type',
    'Attribute Name 3': 'Nombre de dents',
    'Attribute Name 4': 'Force d\'éjection (N)',
    'Attribute Name 5': 'Poids',
    'Attribute Name 6': 'Fixation de colonne de direction',
    'Attribute Name 7': 'Diamètre du disque',
    'Attribute Name 8': 'Info complementaire 1',
    'Attribute Name 9': 'OE',
    # Ajoutez d'autres mappings de header si nécessaire
}

# Mapping spécifique pour le fichier "valeo"
valeo_mip_mapping = {
    'MPN (or OE)': 'MPN',
    'Attribute Name 1': 'Brand',
    # Ajoutez d'autres mappings de champ spécifiques si nécessaire
}

# Dictionnaire des clés de fusion de chaque fichier d'input
fusion_keys = {
    'test.csv': 'clé de fuuuuusion',
    'valeo.csv': 'MPN (or OE)',
    # Ajoutez d'autres clés de fusion pour les autres fichiers d'input
}

# Conversion des fichiers xlsx en csv
for input_file in os.listdir(input_folder):
    if input_file.endswith('.xlsx') and input_file != 'TecDoc.xlsx':
        input_path = os.path.join(input_folder, input_file)
        converted_file = os.path.splitext(input_file)[0] + '.csv'
        converted_path = os.path.join(converted_folder, converted_file)

        # Charger le fichier xlsx
        wb = load_workbook(input_path)
        sheet = wb.active

        # Écrire le contenu dans un fichier csv
        with open(converted_path, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Le fichier '{input_file}' a été converti en CSV : {converted_file}")

# Conversion du fichier TecDoc Excel en CSV
tecdoc_path = os.path.join('ExempleTD', 'TecDoc.xlsx')
converted_tecdoc_file = 'TecDoc.csv'
converted_tecdoc_path = os.path.join(converted_tecdoc_folder, converted_tecdoc_file)

# Charger le fichier TecDoc Excel
tecdoc_wb = load_workbook(tecdoc_path)
tecdoc_sheet = tecdoc_wb.active

# Écrire le contenu dans un fichier csv en utilisant le mapping générique pour les fichiers MIP
with open(converted_tecdoc_path, 'w', newline='', encoding='utf-8') as tecdoc_csv_file:
    tecdoc_writer = csv.writer(tecdoc_csv_file)
    for row in tecdoc_sheet.iter_rows(values_only=True):
        tecdoc_writer.writerow(row)

print(f"Le fichier TecDoc Excel a été converti en CSV : {converted_tecdoc_file}")

# Conversion du fichier MIP.xlsx en MIP.csv
mip_input_path = os.path.join('MIP', 'MIP.xlsx')
converted_mip_file = 'MIP.csv'
converted_mip_folder = 'MIP_converted'  # Dossier pour les fichiers MIP convertis
converted_mip_path = os.path.join(converted_mip_folder, converted_mip_file)

# Créer le dossier MIP_converted s'il n'existe pas déjà
os.makedirs(converted_mip_folder, exist_ok=True)

# Charger le fichier MIP.xlsx
mip_wb = load_workbook(mip_input_path)
mip_sheet = mip_wb.active

# Écrire le contenu dans un fichier CSV
with open(converted_mip_path, 'w', newline='', encoding='utf-8') as mip_csv_file:
    mip_writer = csv.writer(mip_csv_file)
    for row in mip_sheet.iter_rows(values_only=True):
        mip_writer.writerow(row)

print(f"Le fichier MIP.xlsx a été converti en CSV : {converted_mip_file}")

# Traitement des fichiers de fusion
for fusion_file in os.listdir(fusion_folder):
    if fusion_file.endswith('.csv'):
        fusion_path = os.path.join(fusion_folder, fusion_file)

        # Lire le fichier de fusion
        with open(fusion_path, 'r', encoding='utf-8') as fusion_csv_file:
            fusion_reader = csv.DictReader(fusion_csv_file)

            # Récupérer le nom du fichier d'input
            input_file_name = fusion_file.replace('fusion-', '').replace('-tecdoc.csv', '')

            # Vérifier si le fichier de fusion a un mapping spécifique
            if input_file_name == 'valeo':
                current_header_mapping = valeo_mip_mapping
            else:
                current_header_mapping = header_mapping

            # Récupérer la clé de fusion pour le fichier d'input
            fusion_key = fusion_keys.get(fusion_file, 'MPN')

            # Créer un nouveau fichier MIP pour la fusion
            mip_file = f"MIP-{input_file_name}.csv"
            mip_path = os.path.join(mip_folder, mip_file)

            # Écrire les en-têtes MIP génériques pour les fichiers de fusion
            with open(mip_path, 'w', newline='', encoding='utf-8') as mip_csv_file:
                mip_writer = csv.writer(mip_csv_file)
                mip_writer.writerow(list(mip_mapping.values()))

                for row in fusion_reader:
                    mip_row = []
                    for header in list(mip_mapping.keys()):
                        if header == 'MPN':
                            mip_row.append(row.get(fusion_key, ''))
                        else:
                            mapped_header = current_header_mapping.get(header, '')
                            mip_row.append(row.get(mapped_header, ''))

                    mip_writer.writerow(mip_row)

            print(f"Le fichier de fusion '{fusion_file}' a été traité et un nouveau fichier MIP a été créé : {mip_file}")
