import os
import csv
from openpyxl import load_workbook

# Dossier d'entrée et de sortie
input_folder = 'input'
tecdoc_folder = "ExempleTD"
converted_folder = 'converted_input_files'
fusion_folder = 'fusion'
mip_folder = 'MIP_done'
converted_tecdoc_folder = 'converted_tecdoc_files'
mip_fusion_folder = 'MIP_fusion'

# Créer les dossiers de sortie si nécessaire
os.makedirs(converted_folder, exist_ok=True)
os.makedirs(fusion_folder, exist_ok=True)
os.makedirs(mip_folder, exist_ok=True)
os.makedirs(converted_tecdoc_folder, exist_ok=True)
os.makedirs(mip_fusion_folder, exist_ok=True)

# Mapping générique pour les fichiers MIP
mip_mapping = {
    'MPN': 'MPN',
    'Attribute Value 1': 'Marque',
    'Attribute Value 2': 'Type',
    'Attribute Value 3': 'Nombre de dents',
    'Attribute Value 4': 'Force d\'éjection (N)',
    'Attribute Value 5': 'Poids',
    'Attribute Value 6': 'Fixation de colonne de direction',
    'Attribute Value 7': 'Diamètre du disque',
    'Attribute Value 8': 'Info complementaire 1',
    'Attribute Value 9': 'OE',
}

# Mapping générique pour écrire le header des champs des fichiers fusion dans les valeur des champs MIP correspondant
header_mapping = {
    'Attribute Name 1': 'Marque',
    'Attribute Name 2': 'Type',
    'Attribute Name 3': 'Nombre de dents',
    'Attribute Name 4': 'Force d\'éjection (N)',
    'Attribute Name 5': 'Poids',
    'Attribute Name 6': 'Fixation de colonne de direction',
    'Attribute Name 7': 'Diamètre du disque',
    'Attribute Name 8': 'Info complementaire 1',
    'Attribute Name 9': 'OE',
    # Ajoutez d'autres mappings de header si nécessaire
}

# Mapping spécifique pour le fichier "valeo"
valeo_mip_mapping = {
    'MPN (or OE)': 'MPN',
    'Attribute Name 1': 'Brand',
    # Ajoutez d'autres mappings de champ spécifiques si nécessaire
}

# Dictionnaire des clés de fusion de chaque fichier d'input
fusion_keys = {
    'test.csv': 'code de fuuuuusion',
    'valeo.csv': 'MPN (or OE)',
    # Ajoutez d'autres clés de fusion pour les autres fichiers d'input
}

# Conversion des fichiers xlsx en csv
for input_file in os.listdir(input_folder):
    if input_file.endswith('.xlsx') and input_file != 'TecDoc.xlsx':
        input_path = os.path.join(input_folder, input_file)
        converted_file = os.path.splitext(input_file)[0] + '.csv'
        converted_path = os.path.join(converted_folder, converted_file)

        # Charger le fichier xlsx
        wb = load_workbook(input_path)
        sheet = wb.active

        # Lire les données et les écrire dans le fichier csv
        with open(converted_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Le fichier '{input_file}' a été converti en CSV : {converted_file}")

# Conversion du fichier TecDoc
tecdoc_file = 'TecDoc.xlsx'
tecdoc_path = os.path.join(tecdoc_folder, tecdoc_file)
converted_tecdoc_file = 'TecDoc.csv'
converted_tecdoc_path = os.path.join(converted_tecdoc_folder, converted_tecdoc_file)

# Charger le fichier xlsx TecDoc
tecdoc_wb = load_workbook(tecdoc_path)
tecdoc_sheet = tecdoc_wb.active

# Lire les en-têtes du fichier TecDoc
tecdoc_headers = [cell.value for cell in tecdoc_sheet[1]]

# Lire les données et les écrire dans le fichier csv TecDoc
with open(converted_tecdoc_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(tecdoc_headers)
    for row in tecdoc_sheet.iter_rows(min_row=2, values_only=True):
        writer.writerow(row)

print(f"Le fichier '{tecdoc_file}' a été converti en CSV : {converted_tecdoc_file}")

# Fusion des fichiers d'input avec TecDoc
for input_file in os.listdir(converted_folder):
    if input_file.endswith('.csv'):
        input_path = os.path.join(converted_folder, input_file)
        fusion_file = 'fusion-' + input_file.split('.')[0] + '-tecdoc.csv'
        fusion_path = os.path.join(fusion_folder, fusion_file)
        

        # Charger les fichiers d'input et TecDoc
        with open(input_path, 'r', encoding='utf-8') as input_csv, \
                open(converted_tecdoc_path, 'r', encoding='utf-8') as tecdoc_csv, \
                open(fusion_path, 'w', newline='', encoding='utf-8') as fusion_csv, \

            # Lire les données des fichiers d'input et TecDoc
            input_reader = csv.DictReader(input_csv)
            tecdoc_reader = csv.DictReader(tecdoc_csv)

            # Utiliser DictWriter pour écrire les données fusionnées
            fieldnames = input_reader.fieldnames + tecdoc_reader.fieldnames[1:]  # Exclure la première colonne de TecDoc
            fusion_writer = csv.DictWriter(fusion_csv, fieldnames=fieldnames)
            mip_fieldnames = ['MPN', 'Marque', 'Type', 'Nombre de dents', 'Force d\'éjection (N)', 'Poids',
                              'Fixation de colonne de direction', 'Diamètre du disque', 'Info complementaire 1',
                              'OE']
            mip_fusion_writer = csv.DictWriter(mip_fusion_csv, fieldnames=mip_fieldnames)
            fusion_writer.writeheader()
            mip_fusion_writer.writeheader()

            # Créer un dictionnaire des données TecDoc indexé par la clé de fusion
            tecdoc_data = {}
            for tecdoc_row in tecdoc_reader:
                key = tecdoc_row['MPN']
                tecdoc_data[key] = tecdoc_row

            # Effectuer la fusion des données
            for input_row in input_reader:
                fusion_row = {}

                # Parcourir les champs de fusion et effectuer le mapping
                for fusion_field, mip_field in header_mapping.items():
                    fusion_value = input_row.get(fusion_field)
                    mip_field = mip_mapping.get(mip_field, mip_field)
                    fusion_row[mip_field] = fusion_value

                # Rechercher la correspondance dans les données TecDoc
                fusion_key = fusion_keys.get(input_file)
                input_key = fusion_row.get(fusion_key)
                tecdoc_row = tecdoc_data.get(input_key)

                if tecdoc_row:
                    # Fusionner les données TecDoc avec les données d'input
                    fusion_row.update(tecdoc_row)

                fusion_writer.writerow(fusion_row)
                mip_fusion_writer.writerow(fusion_row)

        print(f"Le fichier '{input_file}' a été fusionné avec TecDoc : {fusion_file}")

# Écriture des données fusionnées dans les fichiers MIP-fusion
for fusion_file in os.listdir(fusion_folder):
    if fusion_file.endswith('-tecdoc.csv'):
        fusion_path = os.path.join(fusion_folder, fusion_file)
        mip_fusion_file = 'MIP-' + fusion_file.split('.')[0] + '.csv'
        mip_fusion_path = os.path.join(mip_fusion_folder, mip_fusion_file)

        # Charger les fichiers fusion et MIP
        with open(fusion_path, 'r', encoding='utf-8') as fusion_csv, \
                open(converted_tecdoc_path, 'r', encoding='utf-8') as mip_csv, \
                open(mip_fusion_path, 'w', newline='', encoding='utf-8') as mip_fusion_csv:

            # Lire les données des fichiers fusion et MIP
            fusion_reader = csv.DictReader(fusion_csv)
            mip_reader = csv.DictReader(mip_csv)

            # Utiliser DictWriter pour écrire les données fusionnées
            mip_fieldnames = mip_reader.fieldnames
            mip_fusion_writer = csv.DictWriter(mip_fusion_csv, fieldnames=mip_fieldnames)
            mip_fusion_writer.writeheader()

            # Créer un dictionnaire des données fusion indexé par la clé de fusion
            fusion_data = {}
            for fusion_row in fusion_reader:
                key = fusion_row['MPN']
                fusion_data[key] = fusion_row

            # Effectuer l'écriture des données fusionnées dans les fichiers MIP-fusion
            for mip_row in mip_reader:
                mip_key = mip_row['MPN']

                # Rechercher la correspondance dans les données fusion
                if mip_key in fusion_data:
                    fusion_row = fusion_data[mip_key]
                    mip_fusion_writer.writerow(fusion_row)

        print(f"Les données fusionnées du fichier '{fusion_file}' ont été écrites dans : {mip_fusion_file}")
